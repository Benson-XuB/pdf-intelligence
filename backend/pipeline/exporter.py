import re
from pathlib import Path
from typing import Union

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from backend.pipeline.models import FinalTable

# ── colour palette ──────────────────────────────────────────────
HEADER_BG = "1F4E79"       # dark blue
HEADER_FG = "FFFFFF"       # white text
EVEN_ROW = "F2F7FB"        # pale blue-grey
BORDER_COLOR = "BFBFBF"    # soft grey
TITLE_COLOR = "1F4E79"
LOW_CONF_BG = "FFF2CC"     # pale yellow (less harsh than bright yellow)
LOW_CONF_FG = "CC0000"     # red text for low-confidence flag

# ── borders ─────────────────────────────────────────────────────
THIN = Side(style="thin", color=BORDER_COLOR)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BOTTOM = Side(style="medium", color=HEADER_BG)
HEADER_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=HEADER_BOTTOM)

# ── fills ───────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
EVEN_FILL = PatternFill(start_color=EVEN_ROW, end_color=EVEN_ROW, fill_type="solid")
LOW_CONF_FILL = PatternFill(start_color=LOW_CONF_BG, end_color=LOW_CONF_BG, fill_type="solid")

# ── fonts ───────────────────────────────────────────────────────
TITLE_FONT = Font(name="Calibri", bold=True, size=13, color=TITLE_COLOR)
SUBTITLE_FONT = Font(name="Calibri", size=10, color="808080")
HEADER_FONT = Font(name="Calibri", bold=True, size=10, color=HEADER_FG)
DATA_FONT = Font(name="Calibri", size=10)
NUMBER_FONT = Font(name="Calibri", size=10)

# ── alignments ──────────────────────────────────────────────────
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TEXT_ALIGN = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGN = Alignment(horizontal="right", vertical="center")

# ── regex for smart value parsing ───────────────────────────────
_NUM_RE = re.compile(
    r"^\s*"
    r"(?P<neg>\(?\s*)?"
    r"(?P<sym>[$€¥£])?\s*"
    r"(?P<num>[\d,]+(?:\.\d+)?)"
    r"\s*(?P<close>\))?"
    r"\s*$"
)


def _smart_value(raw: object) -> Union[str, float, int]:
    """Try to parse a cell as a number; fall back to cleaned string."""
    if raw is None:
        return ""
    text = str(raw).strip()
    if not text or text in ("-", "—", "–", "\u2212", "N/A", "n/a"):
        return text

    m = _NUM_RE.match(text)
    if not m:
        return text

    num_str = m.group("num").replace(",", "")
    try:
        value = float(num_str)
    except ValueError:
        return text

    if m.group("neg") and "(" in m.group("neg"):
        value = -value
    elif text.startswith("-"):
        value = -value

    if value == int(value) and "." not in num_str:
        return int(value)
    return value


def _apply_number_format(cell, val) -> None:
    """Apply appropriate number format based on magnitude."""
    if not isinstance(val, (int, float)):
        return
    abs_val = abs(val)
    if abs_val >= 1_000_000_000:
        cell.number_format = "#,##0,,.00"  # billions with decimal
    elif abs_val >= 1_000_000:
        cell.number_format = "#,##0,.00"   # millions with decimal
    elif abs_val >= 1_000:
        cell.number_format = "#,##0"       # thousands
    else:
        cell.number_format = "#,##0.00" if isinstance(val, float) else "#,##0"


def _auto_width(ws, header_start_row: int = 4, min_width: int = 10, max_width: int = 45) -> None:
    """Auto-fit column widths with CJK-aware sizing."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            if cell.row < header_start_row:
                continue  # skip title rows
            val = str(cell.value) if cell.value is not None else ""
            # Rough CJK width: each CJK char ≈ 2.2 Latin chars
            cjk = sum(1 for c in val if ord(c) > 0x2FFF)
            latin = len(val) - cjk
            lengths.append(latin + cjk * 2.2)
        best = max(lengths) if lengths else min_width
        best = max(best + 3, min_width)
        best = min(best, max_width)
        ws.column_dimensions[col_letter].width = best


# ── Main Exporter ───────────────────────────────────────────────
class ExcelExporter:
    def export(
        self,
        tables: list[FinalTable],
        metadata: dict,
        output_path: str,
        low_confidence_threshold: float = 0.85,
    ) -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)

        # ── Index sheet ──────────────────────────────────────────
        self._write_index(wb, tables, metadata)

        # ── Table sheets ─────────────────────────────────────────
        for idx, table in enumerate(tables):
            sheet_name = f"Table {idx + 1} — Pg {table.page_num + 1}"
            ws = wb.create_sheet(sheet_name[:31])
            self._write_table_sheet(ws, table, idx, low_confidence_threshold)

        wb.save(output_path)
        return output_path

    # ── Index sheet ──────────────────────────────────────────────
    def _write_index(self, wb: Workbook, tables: list[FinalTable], metadata: dict) -> None:
        ws = wb.create_sheet("Index", 0)

        # Title
        ws.merge_cells("A1:G1")
        title = ws.cell(row=1, column=1, value="PDF Intelligence — Extraction Report")
        title.font = Font(name="Calibri", bold=True, size=16, color=HEADER_BG)
        title.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 32

        # Metadata section
        row = 3
        ws.cell(row=row, column=1, value="File").font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=2, value=metadata.get("source_file", "—")).font = SUBTITLE_FONT
        row += 1
        ws.cell(row=row, column=1, value="Pages").font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=2, value=str(metadata.get("total_pages", "—"))).font = SUBTITLE_FONT
        row += 1
        ws.cell(row=row, column=1, value="Tables Found").font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=2, value=str(metadata.get("table_count", len(tables)))).font = SUBTITLE_FONT

        # Table summary heading
        row += 2
        summary_headers = ["#", "Sheet", "Source", "Page", "Confidence", "Rows × Cols", ""]
        for ci, h in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = HEADER_BORDER
        ws.row_dimensions[row].height = 22

        # Table summary rows
        for tidx, table in enumerate(tables):
            row += 1
            df = table.dataframe
            vals = [
                tidx + 1,
                f"Table {tidx + 1} — Pg {table.page_num + 1}",
                table.source,
                table.page_num + 1,
                f"{table.confidence:.0%}",
                f"{len(df)} × {len(df.columns)}",
                "",
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = DATA_FONT
                cell.border = CELL_BORDER
                cell.alignment = TEXT_ALIGN if ci <= 4 else NUMBER_ALIGN
                if row % 2 == 0:
                    cell.fill = EVEN_FILL

        _auto_width(ws, header_start_row=1)
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 30
        ws.freeze_panes = "A8"

    # ── Per-table sheet ──────────────────────────────────────────
    def _write_table_sheet(
        self, ws, table: FinalTable, idx: int, low_threshold: float
    ) -> None:
        df = table.dataframe
        ncols = len(df.columns)
        is_low_conf = table.confidence < low_threshold

        # Row 1: Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        title_text = f"Table {idx + 1} — Page {table.page_num + 1}"
        title = ws.cell(row=1, column=1, value=title_text)
        title.font = TITLE_FONT
        title.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        # Row 2: Metadata line
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        meta_text = f"Source: {table.source}    |    Confidence: {table.confidence:.1%}"
        if is_low_conf:
            meta_text += "    ⚠ LOW CONFIDENCE — manual review recommended"
        meta = ws.cell(row=2, column=1, value=meta_text)
        meta.font = Font(name="Calibri", size=10, color=LOW_CONF_FG if is_low_conf else "808080")
        meta.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20

        # Row 3: empty spacer
        ws.row_dimensions[3].height = 4

        # Row 4: Header row
        header_row = 4
        for ci, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=header_row, column=ci, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = HEADER_BORDER
        ws.row_dimensions[header_row].height = 24

        # Rows 5+: Data
        for ri, row_tuple in enumerate(df.itertuples(index=False)):
            excel_row = header_row + 1 + ri
            data_row: list = []
            for ci, raw_val in enumerate(row_tuple, 1):
                cell_val = _smart_value(raw_val)
                is_number = isinstance(cell_val, (int, float))
                cell = ws.cell(row=excel_row, column=ci, value=cell_val)
                cell.font = DATA_FONT
                cell.border = CELL_BORDER
                cell.alignment = NUMBER_ALIGN if is_number else TEXT_ALIGN

                if is_number:
                    _apply_number_format(cell, cell_val)

                data_row.append(cell_val)
                first_data_column = 1

        # Alternating row colors for data rows
        data_start = header_row + 1
        data_end = header_row + len(df)
        for r in range(data_start, data_end + 1):
            if (r - data_start) % 2 == 1:
                for c in range(1, ncols + 1):
                    cell = ws.cell(row=r, column=c)
                    if not cell.fill or cell.fill.start_color.rgb == "00000000":
                        cell.fill = EVEN_FILL

        # Low-confidence annotation
        if is_low_conf:
            for ci in range(1, ncols + 1):
                hdr = ws.cell(row=header_row, column=ci)
                hdr.fill = LOW_CONF_FILL
                hdr.font = Font(name="Calibri", bold=True, size=10, color=LOW_CONF_FG)
            note = ws.cell(row=header_row, column=1)
            note.comment = Comment("Low confidence — manual review recommended", "pdf-intelligence")

        # Column widths
        _auto_width(ws, header_start_row=header_row)

        # Freeze header row
        ws.freeze_panes = f"A{header_row + 1}"

        # Auto-filter
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{header_row}"
