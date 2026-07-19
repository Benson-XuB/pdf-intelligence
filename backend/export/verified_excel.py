from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.global_schema.models import CompanyFinancials, StatementType
from backend.global_schema.registry import GLOBAL_FIELDS_V1
from backend.markets.us.filing_resolver import FilingDocument
from backend.validation.reconciliation import (
    MatchStatus,
    ReconciliationItem,
    ReconciliationReport,
    TrustLevel,
)

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")


class VerifiedExcelExporter:
    def export(
        self,
        xbrl: CompanyFinancials,
        pdf: CompanyFinancials,
        reconciliation: ReconciliationReport,
        filing: FilingDocument,
        output_path: str,
    ) -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)

        self._write_dashboard(wb, xbrl, reconciliation, filing)
        self._write_reconciliation_sheet(wb, reconciliation)
        self._write_authoritative_sheet(wb, reconciliation)

        for statement, title in (
            (StatementType.INCOME, "利润表"),
            (StatementType.BALANCE, "资产负债表"),
            (StatementType.CASHFLOW, "现金流量表"),
        ):
            self._write_statement_detail(
                wb,
                title=title,
                statement=statement,
                reconciliation=reconciliation,
            )

        wb.save(output_path)
        return output_path

    def _style_header_row(self, ws, row: int, col_count: int) -> None:
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width(self, ws, max_width: int = 28) -> None:
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            width = min(
                max_width,
                max(len(str(cell.value or "")) for cell in col_cells) + 2,
            )
            ws.column_dimensions[letter].width = width

    def _write_dashboard(
        self,
        wb: Workbook,
        xbrl: CompanyFinancials,
        report: ReconciliationReport,
        filing: FilingDocument,
    ) -> None:
        ws = wb.create_sheet("产品总览", 0)
        ws["A1"] = "Verified Financial Intelligence"
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:D1")

        readiness = "可用于生产分析" if report.mismatch_count == 0 and report.verification_rate >= 0.9 else "建议人工复核"
        rows = [
            ("Ticker", xbrl.ticker),
            ("Company", xbrl.company_name),
            ("Market", xbrl.market),
            ("CIK", xbrl.cik),
            ("Standard", xbrl.standard),
            ("Filing Form", filing.form),
            ("Filing Date", filing.filing_date or "local cache"),
            ("Source Document", filing.local_path or ""),
            ("Document Type", report.pdf_source_type),
            ("Periods", ", ".join(report.periods)),
            ("Trust Score", f"{report.trust_score:.1%}"),
            ("Verification Rate", f"{report.verification_rate:.1%}"),
            ("PDF Coverage", f"{report.pdf_coverage_rate:.1%}"),
            ("Matched Checks", str(report.matched_count)),
            ("Mismatch Checks", str(report.mismatch_count)),
            ("XBRL-only Fields", str(len([i for i in report.items if i.status == MatchStatus.XBRL_ONLY]))),
            ("PDF-only Fields", str(len([i for i in report.items if i.status == MatchStatus.PDF_ONLY]))),
            ("Production Readiness", readiness),
        ]
        start = 3
        for idx, (key, value) in enumerate(rows, start):
            ws.cell(row=idx, column=1, value=key).font = BOLD
            ws.cell(row=idx, column=2, value=value)
            if key == "Production Readiness":
                cell = ws.cell(row=idx, column=2)
                cell.fill = GREEN if readiness.startswith("可用") else YELLOW

        ws.cell(row=start + len(rows) + 1, column=1, value="方法说明").font = BOLD
        ws.cell(
            row=start + len(rows) + 2,
            column=1,
            value="官方 XBRL 为主数据源，PDF 文本层独立提取后对账。绿色=校验通过，黄色=仅单一来源，红色=不一致。",
        )
        ws.merge_cells(start_row=start + len(rows) + 2, start_column=1, end_row=start + len(rows) + 2, end_column=4)
        self._auto_width(ws)

    def _write_reconciliation_sheet(self, wb: Workbook, report: ReconciliationReport) -> None:
        ws = wb.create_sheet("校验报告")
        headers = [
            "field_id",
            "科目(EN)",
            "科目(ZH)",
            "period_end",
            "status",
            "trust",
            "xbrl",
            "pdf",
            "delta",
            "delta_pct",
            "authoritative",
            "source",
            "xbrl_tag",
            "pdf_label",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header_row(ws, 1, len(headers))

        for row_idx, item in enumerate(report.items, 2):
            values = [
                item.field_id,
                item.label_en,
                item.label_zh,
                item.period_end,
                item.status.value,
                item.trust_level.value,
                item.xbrl_value,
                item.pdf_value,
                item.delta,
                item.delta_pct,
                item.authoritative_value,
                item.authoritative_source,
                item.xbrl_tag,
                item.pdf_label,
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                if col == 5:
                    cell.fill = self._status_fill(item.status)

        ws.freeze_panes = "A2"
        self._auto_width(ws)

    def _write_authoritative_sheet(self, wb: Workbook, report: ReconciliationReport) -> None:
        ws = wb.create_sheet("推荐数据")
        headers = ["field_id", "科目(EN)", "科目(ZH)", "source", "trust"] + report.periods
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header_row(ws, 1, len(headers))

        lookup = self._pivot_items(report.items)
        for row_idx, field_def in enumerate(GLOBAL_FIELDS_V1, 2):
            ws.cell(row=row_idx, column=1, value=field_def.field_id)
            ws.cell(row=row_idx, column=2, value=field_def.label_en)
            ws.cell(row=row_idx, column=3, value=field_def.label_zh)
            sample = lookup.get(field_def.field_id, {}).get(report.periods[0]) if report.periods else None
            ws.cell(row=row_idx, column=4, value=sample.authoritative_source if sample else "")
            ws.cell(row=row_idx, column=5, value=sample.trust_level.value if sample else "")

            for offset, period in enumerate(report.periods, 6):
                item = lookup.get(field_def.field_id, {}).get(period)
                if not item:
                    continue
                cell = ws.cell(row=row_idx, column=offset, value=item.authoritative_value)
                cell.fill = self._trust_fill(item.trust_level)

        ws.freeze_panes = "F2"
        self._auto_width(ws)

    def _write_statement_detail(
        self,
        wb: Workbook,
        title: str,
        statement: StatementType,
        reconciliation: ReconciliationReport,
    ) -> None:
        ws = wb.create_sheet(title)
        fields = [f for f in GLOBAL_FIELDS_V1 if f.statement == statement]
        lookup = self._pivot_items(reconciliation.items)

        row = 1
        ws.cell(row=row, column=1, value=f"{title} - XBRL vs PDF Reconciliation").font = TITLE_FONT
        row = 3
        headers = ["field_id", "科目", "period", "XBRL", "PDF", "Delta", "Status", "Trust"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        self._style_header_row(ws, row, len(headers))
        row += 1

        for field_def in fields:
            for period in reconciliation.periods:
                item = lookup.get(field_def.field_id, {}).get(period)
                if not item:
                    continue
                ws.cell(row=row, column=1, value=field_def.field_id)
                ws.cell(row=row, column=2, value=f"{field_def.label_en} / {field_def.label_zh}")
                ws.cell(row=row, column=3, value=period)
                ws.cell(row=row, column=4, value=item.xbrl_value)
                ws.cell(row=row, column=5, value=item.pdf_value)
                ws.cell(row=row, column=6, value=item.delta)
                ws.cell(row=row, column=7, value=item.status.value)
                trust_cell = ws.cell(row=row, column=8, value=item.trust_level.value)
                trust_cell.fill = self._trust_fill(item.trust_level)
                status_cell = ws.cell(row=row, column=7)
                status_cell.fill = self._status_fill(item.status)
                row += 1

        ws.freeze_panes = "A4"
        self._auto_width(ws)

    def _pivot_items(self, items: List[ReconciliationItem]) -> dict:
        table = {}
        for item in items:
            table.setdefault(item.field_id, {})[item.period_end] = item
        return table

    def _status_fill(self, status: MatchStatus) -> PatternFill:
        if status == MatchStatus.MATCHED:
            return GREEN
        if status == MatchStatus.MISMATCH:
            return RED
        if status in (MatchStatus.XBRL_ONLY, MatchStatus.PDF_ONLY):
            return YELLOW
        return PatternFill()

    def _trust_fill(self, trust: TrustLevel) -> PatternFill:
        if trust == TrustLevel.VERIFIED:
            return GREEN
        if trust == TrustLevel.REVIEW_REQUIRED:
            return RED
        if trust == TrustLevel.UNVERIFIED:
            return YELLOW
        return BLUE
