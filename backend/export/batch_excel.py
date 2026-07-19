from __future__ import annotations

from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from backend.services.batch_models import BatchVerifyReport

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


class BatchPortfolioExporter:
    def export(self, report: BatchVerifyReport, output_path: str) -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)

        summary = wb.create_sheet("组合汇总")
        summary["A1"] = "Batch Verification Portfolio"
        summary["A1"].font = Font(size=14, bold=True)
        rows = [
            ("Markets", ", ".join(report.markets)),
            ("Tickers", ", ".join(report.tickers)),
            ("Periods", str(report.periods)),
            ("Success", f"{report.success_count}/{len(report.items)}"),
            ("Production Ready", str(report.production_ready_count)),
            ("Avg Trust Score", f"{report.avg_trust_score:.1%}"),
        ]
        for idx, (k, v) in enumerate(rows, 3):
            summary.cell(row=idx, column=1, value=k)
            summary.cell(row=idx, column=2, value=v)

        detail = wb.create_sheet("校验明细")
        headers = [
            "market",
            "ticker",
            "company",
            "success",
            "trust_score",
            "verification_rate",
            "pdf_coverage_rate",
            "production_ready",
            "matched",
            "mismatch",
            "excel_path",
            "errors",
        ]
        for col, header in enumerate(headers, 1):
            cell = detail.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for row_idx, item in enumerate(report.items, 2):
            values = [
                item.market,
                item.ticker,
                item.company_name,
                "Y" if item.success else "N",
                round(item.trust_score, 4),
                round(item.verification_rate, 4),
                round(item.pdf_coverage_rate, 4),
                "Y" if item.production_ready else "N",
                item.matched_count,
                item.mismatch_count,
                item.excel_path or "",
                "; ".join(item.errors[:3]),
            ]
            for col, value in enumerate(values, 1):
                cell = detail.cell(row=row_idx, column=col, value=value)
                if item.production_ready and col == 8:
                    cell.fill = GREEN
                if item.success and item.mismatch_count > 0 and col == 10:
                    cell.fill = RED

        wb.save(output_path)
        return output_path
