from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from backend.global_schema.models import CompanyFinancials, StatementType
from backend.global_schema.registry import GLOBAL_FIELDS_V1, field_by_id


class GlobalExcelExporter:
    def export(
        self,
        financials: CompanyFinancials,
        output_path: str,
        pdf_verification: Optional[Dict[str, Dict[str, bool]]] = None,
    ) -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)

        self._write_summary(wb, financials)
        registry = field_by_id()
        for statement in (StatementType.INCOME, StatementType.BALANCE, StatementType.CASHFLOW):
            fields = [f for f in GLOBAL_FIELDS_V1 if f.statement == statement]
            if not fields:
                continue
            title = {
                StatementType.INCOME: "利润表",
                StatementType.BALANCE: "资产负债表",
                StatementType.CASHFLOW: "现金流量表",
            }[statement]
            self._write_statement_sheet(
                wb,
                sheet_name=title,
                financials=financials,
                fields=fields,
                registry=registry,
                pdf_verification=pdf_verification or {},
            )

        wb.save(output_path)
        return output_path

    def _write_summary(self, wb: Workbook, financials: CompanyFinancials) -> None:
        ws = wb.create_sheet("汇总")
        rows = [
            ("Ticker", financials.ticker),
            ("Company", financials.company_name),
            ("Market", financials.market),
            ("CIK", financials.cik),
            ("Standard", financials.standard),
            ("Periods", ", ".join(financials.periods)),
            ("Field Count", str(len(financials.values))),
            ("Errors", "; ".join(financials.errors) if financials.errors else ""),
        ]
        for idx, (key, value) in enumerate(rows, 1):
            ws.cell(row=idx, column=1, value=key)
            ws.cell(row=idx, column=2, value=value)

    def _write_statement_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        financials: CompanyFinancials,
        fields: List,
        registry: dict,
        pdf_verification: Dict[str, Dict[str, bool]],
    ) -> None:
        ws = wb.create_sheet(sheet_name[:31])
        periods = financials.periods
        headers = ["field_id", "科目(EN)", "科目(ZH)", "scale", "source", "source_tag"] + periods
        bold = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold

        lookup: Dict[str, Dict[str, object]] = {}
        for item in financials.values:
            lookup.setdefault(item.field_id, {})[item.period_end] = item

        for row_idx, field_def in enumerate(fields, 2):
            ws.cell(row=row_idx, column=1, value=field_def.field_id)
            ws.cell(row=row_idx, column=2, value=field_def.label_en)
            ws.cell(row=row_idx, column=3, value=field_def.label_zh)
            ws.cell(row=row_idx, column=4, value=field_def.scale.value)

            field_values = lookup.get(field_def.field_id, {})
            sample = next(iter(field_values.values()), None)
            ws.cell(row=row_idx, column=5, value=getattr(sample, "source", ""))
            ws.cell(row=row_idx, column=6, value=getattr(sample, "source_tag", ""))

            for col_offset, period in enumerate(periods, 7):
                item = field_values.get(period)
                if item is None or item.value is None:
                    ws.cell(row=row_idx, column=col_offset, value="")
                    continue
                ws.cell(row=row_idx, column=col_offset, value=item.value)
                verified = pdf_verification.get(field_def.field_id, {}).get(period)
                if verified is False:
                    ws.cell(row=row_idx, column=col_offset).font = Font(color="FF0000")
