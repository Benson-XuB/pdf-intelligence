from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.global_schema.models import CompanyFinancials, StatementType
from backend.global_schema.registry import GLOBAL_FIELDS_V1, field_by_id
from backend.markets.us.statement_grid_extractor import StatementGrid, _normalize_label
from backend.validation.identity_models import IdentityReport
from backend.validation.reconciliation import ReconciliationItem, ReconciliationReport

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")
FORMULA_FONT = Font(color="003399")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

STATEMENT_SHEETS = {
    "income": "Model - Income",
    "balance": "Model - Balance",
    "cashflow": "Model - Cash Flow",
}

TOTAL_ROW_PATTERNS: Dict[str, List[re.Pattern]] = {
    "balance": [
        re.compile(r"^total assets$"),
        re.compile(r"^total liabilities(?: and equity)?$"),
        re.compile(r"^total (?:stockholders|shareholders).? equity$"),
        re.compile(r"^total equity$"),
    ],
    "income": [
        re.compile(r"^total revenues?$"),
        re.compile(r"^total net sales$"),
        re.compile(r"^net income(?: attributable)?"),
        re.compile(r"^operating income"),
    ],
    "cashflow": [
        re.compile(r"^net cash (?:provided by|used in|from)"),
        re.compile(r"^cash (?:and cash equivalents )?at end of"),
    ],
}


@dataclass
class _SheetLayout:
    header_row: int
    first_data_row: int
    period_cols: Dict[str, int]
    total_rows: List[int]


class FormulaExcelExporter:
    """Export analyst-ready Excel with live formulas (=SUM, balance checks)."""

    def export(
        self,
        *,
        ticker: str,
        company_name: str,
        standard: str,
        reconciliation: ReconciliationReport,
        identity_report: IdentityReport,
        statement_grids: Optional[Dict[str, StatementGrid]] = None,
        authoritative: Optional[CompanyFinancials] = None,
        output_path: str,
    ) -> str:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)

        summary_layout = self._write_summary_sheet(
            wb,
            ticker=ticker,
            company_name=company_name,
            standard=standard,
            reconciliation=reconciliation,
            authoritative=authoritative,
        )
        self._write_validation_sheet(wb, identity_report, summary_layout)
        if statement_grids:
            for stype, sheet_name in STATEMENT_SHEETS.items():
                grid = statement_grids.get(stype)
                if grid and grid.rows:
                    self._write_statement_model_sheet(wb, sheet_name, grid)

        wb.save(output_path)
        return output_path

    def _write_summary_sheet(
        self,
        wb: Workbook,
        *,
        ticker: str,
        company_name: str,
        standard: str,
        reconciliation: ReconciliationReport,
        authoritative: Optional[CompanyFinancials],
    ) -> _SheetLayout:
        ws = wb.create_sheet("Model Summary", 0)
        ws["A1"] = "Formula-Linked Financial Model"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"{company_name} ({ticker}) · {standard}"
        ws["A3"] = "Authoritative values from verified reconciliation (millions, except EPS)."

        header_row = 5
        headers = ["field_id", "Line Item"] + list(reconciliation.periods)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        lookup = self._pivot_items(reconciliation.items)
        auth_lookup = self._financials_lookup(authoritative) if authoritative else {}
        row = header_row + 1
        total_rows: List[int] = []
        registry = field_by_id()

        for field_def in GLOBAL_FIELDS_V1:
            ws.cell(row=row, column=1, value=field_def.field_id)
            ws.cell(row=row, column=2, value=field_def.label_en)
            has_value = False
            for offset, period in enumerate(reconciliation.periods, 3):
                item = lookup.get(field_def.field_id, {}).get(period)
                value = None
                if item and item.authoritative_value is not None:
                    value = item.authoritative_value
                elif field_def.field_id in auth_lookup.get(period, {}):
                    value = auth_lookup[period][field_def.field_id]
                if value is not None:
                    has_value = True
                    ws.cell(row=row, column=offset, value=value)
            if has_value and field_def.field_id in {
                "total_assets",
                "total_liabilities",
                "total_equity",
            }:
                total_rows.append(row)
            row += 1

        period_cols = {
            period: header_row + 2 + idx for idx, period in enumerate(reconciliation.periods)
        }
        self._auto_width(ws)
        return _SheetLayout(
            header_row=header_row,
            first_data_row=header_row + 1,
            period_cols=period_cols,
            total_rows=total_rows,
        )

    def _write_validation_sheet(
        self,
        wb: Workbook,
        identity_report: IdentityReport,
        summary_layout: _SheetLayout,
    ) -> None:
        ws = wb.create_sheet("Validation")
        ws["A1"] = "Accounting Identity Checks"
        ws["A1"].font = TITLE_FONT

        headers = ["rule_id", "check", "period", "lhs", "rhs", "delta", "status", "formula"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        row = 4
        field_rows = self._summary_field_rows(summary_layout)
        assets_row = field_rows.get("total_assets")
        liabilities_row = field_rows.get("total_liabilities")
        equity_row = field_rows.get("total_equity")

        for item in identity_report.items:
            ws.cell(row=row, column=1, value=item.rule_id)
            ws.cell(row=row, column=2, value=item.label)
            ws.cell(row=row, column=3, value=item.period_end)
            ws.cell(row=row, column=4, value=item.lhs_value)
            ws.cell(row=row, column=5, value=item.rhs_value)
            ws.cell(row=row, column=6, value=item.delta)

            col_letter = get_column_letter(summary_layout.period_cols.get(item.period_end, 3))
            if assets_row and liabilities_row and equity_row and item.rule_id == "balance_equation":
                assets_ref = f"'Model Summary'!{col_letter}{assets_row}"
                liab_ref = f"'Model Summary'!{col_letter}{liabilities_row}"
                equity_ref = f"'Model Summary'!{col_letter}{equity_row}"
                formula = (
                    f'=IF(ABS({assets_ref}-({liab_ref}+{equity_ref}))'
                    f"<MAX(ABS({assets_ref})*0.0001,1),\"PASS\",\"FAIL\")"
                )
                status_cell = ws.cell(row=row, column=7, value=formula)
                status_cell.font = FORMULA_FONT
                ws.cell(row=row, column=8, value=formula)
            else:
                status = "PASS" if item.passed else "FAIL"
                ws.cell(row=row, column=7, value=status)
                ws.cell(row=row, column=7).fill = PASS_FILL if item.passed else FAIL_FILL
            row += 1

        if not identity_report.items and assets_row and liabilities_row and equity_row:
            for period, col_idx in summary_layout.period_cols.items():
                col_letter = get_column_letter(col_idx)
                assets_ref = f"'Model Summary'!{col_letter}{assets_row}"
                liab_ref = f"'Model Summary'!{col_letter}{liabilities_row}"
                equity_ref = f"'Model Summary'!{col_letter}{equity_row}"
                formula = (
                    f'=IF(ABS({assets_ref}-({liab_ref}+{equity_ref}))'
                    f"<MAX(ABS({assets_ref})*0.0001,1),\"PASS\",\"FAIL\")"
                )
                ws.cell(row=row, column=2, value="Total Assets = Liabilities + Equity")
                ws.cell(row=row, column=3, value=period)
                status_cell = ws.cell(row=row, column=7, value=formula)
                status_cell.font = FORMULA_FONT
                row += 1

        self._auto_width(ws)

    @staticmethod
    def _parse_grid_display_amount(text: str) -> Optional[float]:
        """Parse grid cell for model sheet (no minimum magnitude filter)."""
        if not text or not str(text).strip():
            return None
        raw = str(text).strip()
        paren = re.search(r"\(\s*([\d,]+(?:\.\d+)?)\s*\)", raw)
        if paren:
            return -float(paren.group(1).replace(",", ""))
        cleaned = raw.replace("$", "").replace(",", "").strip()
        match = re.search(r"-?[\d]+(?:\.[\d]+)?", cleaned)
        if not match:
            return None
        return float(match.group(0))

    def _write_statement_model_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        grid: StatementGrid,
    ) -> None:
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = sheet_name
        ws["A1"].font = TITLE_FONT
        ws["A2"] = "Line items from filing grid; totals use =SUM(...) over detail rows."

        header_row = 4
        ws.cell(row=header_row, column=1, value="Line Item")
        period_cols: Dict[str, int] = {}
        for idx, period in enumerate(grid.period_ends, 2):
            period_cols[period] = idx
            cell = ws.cell(row=header_row, column=idx, value=period)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        detail_row_indices: List[int] = []
        total_row_indices: List[int] = []
        row = header_row + 1

        for label, raw_values in grid.rows:
            norm = _normalize_label(label)
            if not label.strip() and not any(str(v).strip() for v in raw_values):
                continue
            ws.cell(row=row, column=1, value=label.strip())
            parsed_any = False
            for period, col_idx in period_cols.items():
                period_idx = grid.period_ends.index(period) if period in grid.period_ends else None
                if period_idx is None or period_idx >= len(raw_values):
                    continue
                amount = self._parse_grid_display_amount(raw_values[period_idx])
                if amount is not None:
                    parsed_any = True
                    ws.cell(row=row, column=col_idx, value=amount)
            if parsed_any:
                if self._is_total_row(norm, grid.statement_type):
                    total_row_indices.append(row)
                else:
                    detail_row_indices.append(row)
            row += 1

        if detail_row_indices and total_row_indices:
            start = detail_row_indices[0]
            for total_row in total_row_indices:
                end = total_row - 1
                if end < start:
                    continue
                for col_idx in period_cols.values():
                    col_letter = get_column_letter(col_idx)
                    formula = f"=SUM({col_letter}{start}:{col_letter}{end})"
                    cell = ws.cell(row=total_row, column=col_idx, value=formula)
                    cell.font = FORMULA_FONT
                start = total_row + 1

        ws.freeze_panes = "B5"
        self._auto_width(ws)

    @staticmethod
    def _is_total_row(norm_label: str, statement_type: str) -> bool:
        patterns = TOTAL_ROW_PATTERNS.get(statement_type, [])
        return any(pat.search(norm_label) for pat in patterns)

    @staticmethod
    def _pivot_items(items: Sequence[ReconciliationItem]) -> dict:
        table: dict = {}
        for item in items:
            table.setdefault(item.field_id, {})[item.period_end] = item
        return table

    @staticmethod
    def _financials_lookup(financials: CompanyFinancials) -> Dict[str, Dict[str, float]]:
        table: Dict[str, Dict[str, float]] = {}
        for item in financials.values:
            if item.value is None:
                continue
            table.setdefault(item.period_end, {})[item.field_id] = item.value
        return table

    @staticmethod
    def _summary_field_rows(layout: _SheetLayout) -> Dict[str, int]:
        return {
            "total_assets": layout.first_data_row
            + [f.field_id for f in GLOBAL_FIELDS_V1].index("total_assets"),
            "total_liabilities": layout.first_data_row
            + [f.field_id for f in GLOBAL_FIELDS_V1].index("total_liabilities"),
            "total_equity": layout.first_data_row
            + [f.field_id for f in GLOBAL_FIELDS_V1].index("total_equity"),
        }

    @staticmethod
    def _auto_width(ws, max_width: int = 36) -> None:
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            width = min(
                max_width,
                max(len(str(cell.value or "")) for cell in col_cells) + 2,
            )
            ws.column_dimensions[letter].width = width
