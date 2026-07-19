from __future__ import annotations

import hashlib
import json
import re
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from pydantic import BaseModel

from backend.config import settings
from backend.export.global_excel import GlobalExcelExporter
from backend.markets.us.financials_service import UsFinancialsService
from backend.markets.us.xbrl_adapter import UsSecXbrlAdapter
from backend.pipeline.orchestrator import PipelineOrchestrator, PipelineResult
from backend.auth.database import get_tier_limits, log_usage

FEEDBACK_DIR = Path("data/feedback")
FEEDBACK_DIR.mkdir(parents=True, exist_ok=True)
_FEEDBACK_YELLOW = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

jobs: dict[str, dict] = {}
orchestrator = PipelineOrchestrator()
_jobs_lock = threading.Lock()
_JOB_TIMEOUT_MINUTES = 30
us_xbrl_adapter = UsSecXbrlAdapter()
global_exporter = GlobalExcelExporter()
us_financials_service = UsFinancialsService()


class TableRow(BaseModel):
    cells: list[str]


class TableSummary(BaseModel):
    """"column_index" → {"avg": ..., "sum": ...} for numeric columns."""
    columns: dict[str, dict]


class TableDataResponse(BaseModel):
    table_id: int
    source: str
    page_num: int
    confidence: float
    row_count: int
    col_count: int
    headers: list[str]
    rows: list[TableRow]
    summary: TableSummary
    sheet_index: int = 0  # original index in exporter — maps to table_{sheet_index+1}_pxx sheet


class TablesResponse(BaseModel):
    tables: list[TableDataResponse]


class CellEdit(BaseModel):
    table_id: int
    sheet_index: int = 0
    row: int
    col: int
    original: str
    corrected: str


class FeedbackRequest(BaseModel):
    edits: List[CellEdit]


_NUMERIC_RE = re.compile(r"^[\s$€¥£%\-()]*(\d[\d,.]*)[\s%]*\)?$")


def _parse_number(s: str) -> Optional[float]:
    try:
        m = _NUMERIC_RE.match(s.strip())
        if m:
            num = m.group(1).replace(",", "")
            val = float(num)
            if s.strip().startswith("(") or s.strip().startswith("-"):
                val = -val
            return val
        if s.strip().replace("-", "").isdigit() or s.strip().replace("-", "").replace(".", "").isdigit():
            return float(s.strip().replace(",", ""))
        return None
    except (ValueError, AttributeError):
        return None


def _compute_summary(headers: list[str], rows: list[list[str]]) -> dict:
    col_stats = {}
    for ci, hdr in enumerate(headers):
        nums = []
        for row in rows:
            if ci < len(row):
                n = _parse_number(row[ci])
                if n is not None:
                    nums.append(n)
        if len(nums) >= 2:
            avg = round(sum(nums) / len(nums), 2)
            total = round(sum(nums), 2)
            col_stats[str(ci)] = {"avg": avg, "sum": total, "header": hdr}
    return col_stats


def _tables_to_response(tables: list) -> TablesResponse:
    result = []
    for tid, t in enumerate(tables):
        if hasattr(t, "dataframe"):
            df = t.dataframe
            headers = [str(h) for h in df.columns]
            rows = []
            for _, row in df.iterrows():
                cells = [str(v) if v is not None and str(v) != "nan" else "" for v in row]
                rows.append(TableRow(cells=cells))
            summary_data = _compute_summary(headers, [[c for c in r.cells] for r in rows])
            result.append(TableDataResponse(
                table_id=tid,
                source=t.source if hasattr(t, "source") else "unknown",
                page_num=t.page_num if hasattr(t, "page_num") else 0,
                confidence=t.confidence if hasattr(t, "confidence") else 0.0,
                row_count=len(rows),
                col_count=len(headers),
                headers=headers,
                rows=rows,
                summary=TableSummary(columns=summary_data),
                sheet_index=tid,
            ))
    result.sort(key=lambda t: t.row_count * t.col_count * t.confidence, reverse=True)
    # Build display_id from sorted position, but keep sheet_index from original enumeration
    for display_idx, t in enumerate(result):
        t.table_id = display_idx
    return TablesResponse(tables=result)


class JobStatus(BaseModel):
    job_id: str
    status: str
    filename: str
    qwen_calls: int = 0
    deepseek_calls: int = 0
    total_pages: int = 0
    table_count: int = 0
    errors: list[str] = []
    confidence_summary: dict = {}
    output_ready: bool = False
    created_at: str
    processing_page: int = 0
    processing_stage: str = ""


class GlobalFieldResponse(BaseModel):
    field_id: str
    period_end: str
    fiscal_year: Optional[int]
    value: Optional[float]
    currency: str
    scale: str
    standard: str
    source: str
    source_tag: str
    source_form: str
    filed_date: str


class GlobalFinancialsResponse(BaseModel):
    ticker: str
    company_name: str
    market: str
    cik: str
    standard: str
    periods: list[str]
    values: list[GlobalFieldResponse]
    errors: list[str] = []
    excel_path: Optional[str] = None


class ReconciliationItemResponse(BaseModel):
    field_id: str
    label_en: str
    label_zh: str
    period_end: str
    status: str
    trust_level: str
    xbrl_value: Optional[float]
    pdf_value: Optional[float]
    delta: Optional[float]
    authoritative_value: Optional[float]
    authoritative_source: str


class VerifiedFinancialsResponse(BaseModel):
    ticker: str
    company_name: str
    market: str
    cik: str
    periods: list[str]
    trust_score: float
    verification_rate: float
    pdf_coverage_rate: float = 0.0
    production_ready: bool
    matched_count: int
    mismatch_count: int
    pdf_source: Optional[str]
    pdf_source_type: Optional[str]
    excel_path: Optional[str]
    errors: list[str] = []
    reconciliation: list[ReconciliationItemResponse]


def register(app: FastAPI) -> None:
    @app.get("/api/health")
    def health():
        from backend.api.auth import api_key_manager

        return {
            "status": "ok",
            "qwen_configured": bool(settings.dashscope_api_key),
            "sec_user_agent": settings.sec_user_agent,
            "api_auth_enabled": api_key_manager.auth_enabled,
        }

    @app.get("/api/us/{ticker}/financials", response_model=GlobalFinancialsResponse)
    def get_us_financials(ticker: str, periods: int = 3, export_excel: bool = True):
        if periods < 1 or periods > 5:
            raise HTTPException(400, "Periods must be between 1 and 5")
        try:
            financials = us_xbrl_adapter.fetch(ticker, periods=periods)
        except ValueError as exc:
            raise HTTPException(404, str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(502, str(exc)) from exc

        excel_path = None
        if export_excel:
            output_dir = Path(settings.output_dir) / "global_schema_us"
            output_dir.mkdir(parents=True, exist_ok=True)
            excel_path = str(output_dir / f"{ticker.lower()}_global_schema.xlsx")
            global_exporter.export(financials, excel_path)

        return GlobalFinancialsResponse(
            ticker=financials.ticker,
            company_name=financials.company_name,
            market=financials.market,
            cik=financials.cik,
            standard=financials.standard,
            periods=financials.periods,
            values=[
                GlobalFieldResponse(
                    field_id=v.field_id,
                    period_end=v.period_end,
                    fiscal_year=v.fiscal_year,
                    value=v.value,
                    currency=v.currency,
                    scale=v.scale.value,
                    standard=v.standard,
                    source=v.source,
                    source_tag=v.source_tag,
                    source_form=v.source_form,
                    filed_date=v.filed_date,
                )
                for v in financials.values
            ],
            errors=financials.errors,
            excel_path=excel_path,
        )

    def _verified_response(result) -> VerifiedFinancialsResponse:
        return VerifiedFinancialsResponse(
            ticker=result.ticker,
            company_name=result.company_name,
            market=result.market,
            cik=result.cik,
            periods=result.xbrl.periods,
            trust_score=result.trust_score,
            verification_rate=result.verification_rate,
            pdf_coverage_rate=result.pdf_coverage_rate,
            production_ready=result.is_production_ready,
            matched_count=result.reconciliation.matched_count,
            mismatch_count=result.reconciliation.mismatch_count,
            pdf_source=result.filing.local_path if result.filing else None,
            pdf_source_type=result.reconciliation.pdf_source_type,
            excel_path=result.excel_path,
            errors=result.errors,
            reconciliation=[
                ReconciliationItemResponse(
                    field_id=item.field_id,
                    label_en=item.label_en,
                    label_zh=item.label_zh,
                    period_end=item.period_end,
                    status=item.status.value,
                    trust_level=item.trust_level.value,
                    xbrl_value=item.xbrl_value,
                    pdf_value=item.pdf_value,
                    delta=item.delta,
                    authoritative_value=item.authoritative_value,
                    authoritative_source=item.authoritative_source,
                )
                for item in result.reconciliation.items
            ],
        )

    @app.get("/api/us/{ticker}/verified-financials", response_model=VerifiedFinancialsResponse)
    def get_us_verified_financials(
        ticker: str,
        periods: int = 3,
        document_path: Optional[str] = None,
        export_excel: bool = True,
    ):
        if periods < 1 or periods > 5:
            raise HTTPException(400, "Periods must be between 1 and 5")
        try:
            result = us_financials_service.build_verified_financials(
                ticker=ticker,
                periods=periods,
                document_path=document_path,
                export_excel=export_excel,
            )
        except FileNotFoundError as exc:
            raise HTTPException(404, str(exc)) from exc
        except ValueError as exc:
            raise HTTPException(404, str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(502, str(exc)) from exc
        return _verified_response(result)

    @app.post("/api/us/{ticker}/verified-financials", response_model=VerifiedFinancialsResponse)
    async def post_us_verified_financials(
        ticker: str,
        periods: int = 3,
        export_excel: bool = True,
        file: UploadFile = File(...),
    ):
        if not file.filename:
            raise HTTPException(400, "Please upload a 10-K PDF or HTML filing")
        suffix = Path(file.filename).suffix.lower()
        if suffix not in {".pdf", ".htm", ".html"}:
            raise HTTPException(400, "Only PDF / HTML filing files are supported")

        upload_dir = Path(settings.upload_dir) / "filings"
        upload_dir.mkdir(parents=True, exist_ok=True)
        saved_path = upload_dir / f"{ticker.lower()}_{file.filename}"
        saved_path.write_bytes(await file.read())

        return get_us_verified_financials(
            ticker=ticker,
            periods=periods,
            document_path=str(saved_path),
            export_excel=export_excel,
        )

    @app.get("/api/us/{ticker}/verified-financials/download")
    def download_verified_excel(ticker: str, periods: int = 3):
        if periods < 1 or periods > 5:
            raise HTTPException(400, "Periods must be between 1 and 5")
        try:
            result = us_financials_service.build_verified_financials(
                ticker=ticker,
                periods=periods,
                export_excel=True,
            )
        except (FileNotFoundError, ValueError) as exc:
            raise HTTPException(404, str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(502, str(exc)) from exc

        if not result.excel_path or not Path(result.excel_path).exists():
            raise HTTPException(404, "Excel file not found")
        return FileResponse(
            result.excel_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{ticker.lower()}_verified_financials.xlsx",
        )

    def _update_progress(job_id: str, stage: str, current: int, total: int) -> None:
        with _jobs_lock:
            j = jobs.get(job_id)
            if j:
                j["processing_stage"] = stage
                j["processing_page"] = current
                j["total_pages"] = total

    @app.post("/api/upload", response_model=JobStatus)
    async def upload_pdf(
        request: Request,
        file: UploadFile = File(...),
        enable_docling: Optional[str] = Form(None),
        fast_mode: Optional[str] = Form(None),
        use_vlm: Optional[str] = Form(None),
        use_deepseek: Optional[str] = Form(None),
        confidence_threshold: Optional[str] = Form(None),
    ):
        if not file.filename or not file.filename.lower().endswith(".pdf"):
            raise HTTPException(400, "Only PDF files are supported")

        # Tier enforcement
        user_data = getattr(request.state, "user", None)
        is_guest = getattr(request.state, "is_guest", False)
        tier = user_data.get("tier", "free") if user_data else "free"
        limits = get_tier_limits(tier)

        # File size check
        file_bytes = await file.read()
        file_size_mb = len(file_bytes) / (1024 * 1024)
        if file_size_mb > limits["max_file_size_mb"]:
            raise HTTPException(
                413,
                f"File too large ({file_size_mb:.1f}MB). Max {limits['max_file_size_mb']}MB on {tier} tier. Upgrade to upload larger files.",
            )

        upload_dir = Path(settings.upload_dir)
        upload_dir.mkdir(parents=True, exist_ok=True)
        output_dir = Path(settings.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        job_id = str(uuid.uuid4())
        pdf_path = upload_dir / f"{job_id}_{file.filename}"
        pdf_path.write_bytes(file_bytes)

        # Log usage BEFORE extraction (so middleware can count it for subsequent requests)
        uid = user_data.get("id") if user_data else None
        if uid:
            log_usage(uid, "upload", file.filename)
        elif is_guest:
            from backend.auth.database import record_guest_upload
            from backend.auth.middleware import _client_ip

            record_guest_upload(_client_ip(request))

        # Resolve settings: frontend form values override .env, but tier trumps all
        resolved_enable_docling = enable_docling is not None and enable_docling == "1"
        resolved_fast_mode = fast_mode is not None and fast_mode == "1"
        resolved_use_vlm = use_vlm is not None and use_vlm == "1"
        resolved_threshold = settings.confidence_threshold
        if confidence_threshold is not None:
            try:
                resolved_threshold = float(confidence_threshold)
            except ValueError:
                pass

        # Enforce tier engine limits
        resolved_use_deepseek = use_deepseek is not None and use_deepseek == "1"
        if not limits.get("allow_deepseek", False):
            resolved_use_deepseek = False
        if not limits["allow_docling"]:
            resolved_enable_docling = False
        if not limits["allow_qwen"]:
            resolved_use_vlm = False
            # Also disable Qwen fallback in confidence threshold (lower it so no fallback triggers)
            resolved_threshold = 0.0  # disable Qwen auto-fallback for free tier

        job = {
            "job_id": job_id,
            "status": "processing",
            "filename": file.filename,
            "pdf_path": str(pdf_path),
            "output_path": None,
            "created_at": datetime.utcnow().isoformat(),
            "processing_page": 0,
            "total_pages": 0,
            "processing_stage": "starting",
            "started_at": datetime.utcnow(),
            "user_id": user_data.get("id") if user_data else None,
        }
        with _jobs_lock:
            jobs[job_id] = job

        def _run_extraction() -> None:
            try:
                proc_result: PipelineResult = None

                def _do_process():
                    nonlocal proc_result
                    proc_result = orchestrator.process(
                        str(pdf_path),
                        output_path=str(output_dir / f"{job_id}.xlsx"),
                        progress_callback=lambda stage, current, total: _update_progress(
                            job_id, stage, current, total
                        ),
                        enable_docling=resolved_enable_docling,
                        fast_mode=resolved_fast_mode,
                        use_vlm=resolved_use_vlm,
                        use_deepseek=resolved_use_deepseek,
                        confidence_threshold=resolved_threshold,
                    )

                with ThreadPoolExecutor(max_workers=1) as exec:
                    future = exec.submit(_do_process)
                    future.result(timeout=_JOB_TIMEOUT_MINUTES * 60)
                with _jobs_lock:
                    job.update(
                        {
                            "status": "completed",
                            "output_path": proc_result.output_path,
                            "qwen_calls": proc_result.qwen_calls,
                            "deepseek_calls": proc_result.deepseek_calls,
                            "total_pages": proc_result.total_pages,
                            "table_count": len(proc_result.tables),
                            "tables": proc_result.tables,
                            "errors": proc_result.errors,
                            "confidence_summary": {
                                str(k): {"score": v.score, "needs_qwen": v.needs_qwen, "reasons": v.reasons}
                                for k, v in proc_result.confidence_reports.items()
                            },
                        }
                    )

                # Log engine usage
                uid = job.get("user_id")
                if uid:
                    if proc_result and proc_result.qwen_calls:
                        log_usage(uid, "qwen", file.filename)
                    if proc_result and proc_result.deepseek_calls:
                        log_usage(uid, "deepseek", file.filename)
            except FutureTimeoutError:
                with _jobs_lock:
                    job.update({"status": "timeout", "errors": [f"Processing timed out after {_JOB_TIMEOUT_MINUTES} minutes"]})
            except Exception as exc:
                with _jobs_lock:
                    job.update({"status": "failed", "errors": [str(exc)]})

        threading.Thread(target=_run_extraction, daemon=True).start()

        return JobStatus(
            job_id=job_id,
            status="processing",
            filename=file.filename,
            qwen_calls=0,
            deepseek_calls=0,
            total_pages=0,
            table_count=0,
            errors=[],
            confidence_summary={},
            output_ready=False,
            created_at=job["created_at"],
        )

    @app.get("/api/jobs/{job_id}", response_model=JobStatus)
    def get_job(job_id: str):
        job = jobs.get(job_id)
        if not job:
            raise HTTPException(404, "Job not found")

        # detect stalled processing jobs
        status = job["status"]
        errors = job.get("errors", [])
        if status == "processing":
            started = job.get("started_at")
            if started:
                elapsed = (datetime.utcnow() - started).total_seconds()
                if elapsed > _JOB_TIMEOUT_MINUTES * 60:
                    status = "timeout"
                    errors = [f"Processing timed out after {_JOB_TIMEOUT_MINUTES} minutes"]

        return JobStatus(
            job_id=job_id,
            status=status,
            filename=job["filename"],
            qwen_calls=job.get("qwen_calls", 0),
            deepseek_calls=job.get("deepseek_calls", 0),
            total_pages=job.get("total_pages", 0),
            table_count=job.get("table_count", 0),
            errors=errors,
            confidence_summary=job.get("confidence_summary", {}),
            output_ready=bool(job.get("output_path")),
            created_at=job["created_at"],
            processing_page=job.get("processing_page", 0),
            processing_stage=job.get("processing_stage", ""),
        )

    @app.get("/api/jobs/{job_id}/download")
    def download_excel(job_id: str):
        job = jobs.get(job_id)
        if not job or not job.get("output_path"):
            raise HTTPException(404, "File not ready")
        path = Path(job["output_path"])
        if not path.exists():
            raise HTTPException(404, "Output file not found")

        feedback_path = FEEDBACK_DIR / f"{job_id}.jsonl"
        if feedback_path.exists():
            annotated_path = path.parent / f"{path.stem}_annotated{path.suffix}"
            _write_annotated_excel(path, feedback_path, annotated_path)
            return FileResponse(
                annotated_path,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=f"{Path(job['filename']).stem}.xlsx",
            )

        return FileResponse(
            path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{Path(job['filename']).stem}.xlsx",
        )

    @app.get("/api/jobs/{job_id}/tables", response_model=TablesResponse)
    def get_job_tables(job_id: str):
        job = jobs.get(job_id)
        if not job:
            raise HTTPException(404, "Job not found")
        if job.get("status") != "completed":
            raise HTTPException(400, "Job not yet completed")
        tables = job.get("tables", [])
        return _tables_to_response(tables)

    @app.post("/api/jobs/{job_id}/feedback")
    def save_feedback(job_id: str, body: FeedbackRequest):
        job = jobs.get(job_id)
        if not job:
            raise HTTPException(404, "Job not found")

        pdf_path = job.get("pdf_path", "")
        pdf_hash = ""
        if pdf_path and Path(pdf_path).exists():
            pdf_hash = hashlib.md5(Path(pdf_path).read_bytes()).hexdigest()[:16]

        feedback_path = FEEDBACK_DIR / f"{job_id}.jsonl"
        with open(feedback_path, "w") as f:
            for edit in body.edits:
                record = {
                    "job_id": job_id,
                    "filename": job.get("filename", ""),
                    "pdf_hash": pdf_hash,
                    "table_id": edit.table_id,
                    "row": edit.row,
                    "col": edit.col,
                    "original": edit.original,
                    "corrected": edit.corrected,
                    "timestamp": datetime.utcnow().isoformat(),
                }
                f.write(json.dumps(record, ensure_ascii=False) + "\n")

        return {"saved": len(body.edits), "job_id": job_id}


def _write_annotated_excel(original_path: Path, feedback_path: Path, output_path: Path):
    edits_by_sheet: dict[int, list[dict]] = {}
    with open(feedback_path) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            sid = rec.get("sheet_index", rec["table_id"])
            edits_by_sheet.setdefault(sid, []).append(rec)

    wb = load_workbook(original_path)
    for sheet_idx, edits in edits_by_sheet.items():
        prefix_lower = f"table {sheet_idx + 1} —"
        target_sheet = None
        for ws in wb.worksheets:
            if ws.title.lower().startswith(prefix_lower):
                target_sheet = ws
                break
        if target_sheet is None:
            continue
        for edit in edits:
            # Header row is now at row 4 (title + meta + spacer), data starts at row 5
            excel_row = edit["row"] + 5
            excel_col = edit["col"] + 1
            cell = target_sheet.cell(row=excel_row, column=excel_col)
            cell.fill = _FEEDBACK_YELLOW
            comment_text = f"User correction: {edit['original']} -> {edit['corrected']}"
            cell.comment = Comment(comment_text, "PDF Intelligence")
    wb.save(output_path)
