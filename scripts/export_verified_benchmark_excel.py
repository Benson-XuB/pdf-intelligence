#!/usr/bin/env python3
"""导出 verified benchmark Excel：支持 US 10-K、港股，或合并 40 家。"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Protocol, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from backend.global_schema.registry import GLOBAL_FIELDS_V1, field_by_id
from backend.markets.hk.financials_service import HkFinancialsService
from backend.markets.us.financials_service import UsFinancialsService
from backend.validation.reconciliation import MatchStatus, ReconciliationItem

US_REPORT = ROOT / "tests/benchmark/financial_10k/verified_accuracy_report.json"
HK_REPORT = ROOT / "tests/benchmark/financial_hk/verified_accuracy_report.json"
US_OUTPUT = ROOT / "data/outputs/financial_10k_benchmark.xlsx"
HK_OUTPUT = ROOT / "data/outputs/financial_hk_benchmark.xlsx"
MERGED_OUTPUT = ROOT / "data/outputs/verified_benchmark_merged.xlsx"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
BOLD = Font(bold=True)
HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")


class FinancialsService(Protocol):
    def build_verified_financials(
        self,
        ticker: str,
        periods: int = 3,
        document_path: Optional[str] = None,
        export_excel: bool = False,
    ): ...


def _style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _status_fill(status: MatchStatus) -> PatternFill:
    if status == MatchStatus.MATCHED:
        return GREEN
    if status == MatchStatus.MISMATCH:
        return RED
    if status == MatchStatus.SKIPPED:
        return BLUE
    return YELLOW


def _dedupe_errors(errors: List[str]) -> List[str]:
    seen: set[str] = set()
    out: List[str] = []
    for err in errors:
        if err not in seen:
            seen.add(err)
            out.append(err)
    return out


def _us_templates() -> List[dict]:
    from tests.benchmark.financial_10k.corpus import available_entries

    return [
        {
            "id": e["id"],
            "name": e["name"],
            "pdf": e["pdf"],
            "tier": "A",
            "market": "US",
            "cross_list": e["id"],
        }
        for e in available_entries()
    ]


def _hk_templates() -> List[dict]:
    from tests.benchmark.financial_hk.corpus import available_entries

    return [
        {
            "id": e["id"],
            "name": e["name"],
            "pdf": e["pdf"],
            "tier": e.get("tier", "A"),
            "market": "HK",
            "cross_list": e.get("sec_ticker") or "",
        }
        for e in available_entries()
    ]


def _collect_verified_rows(
    service: FinancialsService,
    templates: List[dict],
    market: str,
    periods: int = 3,
) -> Tuple[List[dict], List[Tuple[str, str, ReconciliationItem]], Dict[str, List[str]]]:
    summaries: List[dict] = []
    all_items: List[Tuple[str, str, ReconciliationItem]] = []
    warnings: Dict[str, List[str]] = {}

    for tpl in templates:
        pdf_path = tpl["pdf"]
        if not Path(pdf_path).exists():
            continue
        ticker = tpl["id"]
        print(f"  [{market}] {ticker} ...")
        result = service.build_verified_financials(
            ticker=ticker,
            periods=periods,
            document_path=pdf_path,
            export_excel=False,
        )
        summaries.append(
            {
                "id": ticker,
                "name": tpl.get("name", ticker),
                "tier": tpl.get("tier", "A"),
                "market": market,
                "cross_list": tpl.get("cross_list", ""),
                "pdf_extraction_accuracy": getattr(result, "pdf_extraction_accuracy", 1.0),
                "verification_rate": result.verification_rate,
                "trust_score": result.trust_score,
                "pdf_coverage_rate": result.pdf_coverage_rate,
                "production_ready": result.is_production_ready,
                "recon_matched": result.reconciliation.matched_count,
                "recon_mismatch": result.reconciliation.mismatch_count,
                "periods": result.reconciliation.periods,
            }
        )
        for item in result.reconciliation.items:
            all_items.append((market, ticker, item))
        warnings[f"{market}:{ticker}"] = _dedupe_errors(result.errors)

    return summaries, all_items, warnings


def _summaries_from_report(report_path: Path, market: str) -> Tuple[List[dict], Dict[str, List[str]]]:
    report = json.loads(report_path.read_text(encoding="utf-8"))
    summaries: List[dict] = []
    warnings: Dict[str, List[str]] = {}
    for co in report.get("companies", []):
        summaries.append(
            {
                "id": co["id"],
                "name": co.get("name", co["id"]),
                "tier": co.get("tier", "A"),
                "market": market,
                "cross_list": co.get("cross_list_us") or "",
                "pdf_extraction_accuracy": co.get("pdf_extraction_accuracy", 1.0),
                "verification_rate": co["verification_rate"],
                "trust_score": co["trust_score"],
                "pdf_coverage_rate": co.get("pdf_coverage_rate", 0),
                "production_ready": co.get("production_ready", False),
                "recon_matched": co.get("recon_matched", 0),
                "recon_mismatch": co.get("recon_mismatch", 0),
                "periods": [],
            }
        )
        warnings[f"{market}:{co['id']}"] = _dedupe_errors(co.get("errors", []))
    return summaries, warnings


def _portfolio_stats(summaries: List[dict]) -> dict:
    if not summaries:
        return {"count": 0, "avg_verify": 0.0, "avg_trust": 0.0, "avg_pdf_coverage": 0.0, "mismatch": 0}
    return {
        "count": len(summaries),
        "avg_verify": sum(s["verification_rate"] for s in summaries) / len(summaries),
        "avg_trust": sum(s["trust_score"] for s in summaries) / len(summaries),
        "avg_pdf_coverage": sum(s.get("pdf_coverage_rate", 0) for s in summaries) / len(summaries),
        "mismatch": sum(s.get("recon_mismatch", 0) for s in summaries),
    }


def _auto_width(sheet, max_width: int = 40) -> None:
    for col_idx in range(1, sheet.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 12
        for row_idx in range(1, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            width = max(width, len(str(val or "")) + 2)
        sheet.column_dimensions[letter].width = min(width, max_width)


def _write_overview_sheet(ws, summaries: List[dict]) -> None:
    ws.append(["Verified Financials Benchmark — 合并报告"])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(size=14, bold=True, color="1F4E78")
    ws.append(["生成时间", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    ws.append([])

    us = [s for s in summaries if s["market"] == "US"]
    hk = [s for s in summaries if s["market"] == "HK"]
    us_stats = _portfolio_stats(us)
    hk_stats = _portfolio_stats(hk)
    all_stats = _portfolio_stats(summaries)

    ws.append(["市场", "公司数", "平均校验通过率", "平均信任分", "平均PDF覆盖率", "不一致合计"])
    _style_header(ws, ws.max_row, 6)
    for label, stats in (("US 10-K", us_stats), ("港股", hk_stats), ("合计", all_stats)):
        ws.append([
            label,
            stats["count"],
            stats["avg_verify"],
            stats["avg_trust"],
            stats["avg_pdf_coverage"],
            stats["mismatch"],
        ])
        row = ws.max_row
        ws.cell(row, 3).fill = GREEN if stats["avg_verify"] >= 1.0 else YELLOW
        ws.cell(row, 4).fill = GREEN if stats["avg_trust"] >= 0.9 else YELLOW
        ws.cell(row, 5).fill = GREEN if stats["avg_pdf_coverage"] >= 0.9 else YELLOW

    ws.append([])
    ws.append(["说明", "绿色=指标达标；校验数据 sheet 含 PDF×XBRL 对账明细；提取警告仅为诊断信息"])
    ws.merge_cells(f"A{ws.max_row}:F{ws.max_row}")
    ws[f"A{ws.max_row}"].font = Font(italic=True, color="666666")


def _write_summary_sheet(ws, summaries: List[dict]) -> None:
    ws.append(["说明", "本文件「校验数据」含实际财务数值；「提取警告」仅为辅助诊断，不等于校验失败"])
    ws.merge_cells("A1:L1")
    ws["A1"].font = Font(italic=True, color="666666")
    ws.append([])
    headers = [
        "Market", "Ticker", "公司", "Tier", "交叉上市", "XBRL校验", "信任分", "PDF覆盖率", "Ready",
        "对账命中", "不一致", "报告期",
    ]
    ws.append(headers)
    _style_header(ws, ws.max_row, len(headers))
    for s in sorted(summaries, key=lambda x: (x["market"], x["id"])):
        ws.append([
            s["market"],
            s["id"],
            s["name"],
            s.get("tier", ""),
            s.get("cross_list", "") or "—",
            s["verification_rate"],
            s["trust_score"],
            s.get("pdf_coverage_rate", 0),
            "YES" if s.get("production_ready") else "NO",
            s.get("recon_matched", 0),
            s.get("recon_mismatch", 0),
            ", ".join(s.get("periods", [])),
        ])
        row = ws.max_row
        ws.cell(row, 6).fill = GREEN if s["verification_rate"] >= 1.0 else YELLOW
        ws.cell(row, 7).fill = GREEN if s["trust_score"] >= 0.9 else YELLOW
        ws.cell(row, 8).fill = GREEN if s.get("pdf_coverage_rate", 0) >= 0.9 else YELLOW


def _write_recommended_sheet(ws, summaries: List[dict], all_items: List[Tuple[str, str, ReconciliationItem]]) -> None:
    ws.append(["图例", "绿色=PDF与XBRL一致", "蓝色=跳过/不可用", "红色=不一致"])
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(italic=True, color="666666")
    ws.append([])
    registry = field_by_id()
    col = 1
    header_row = 3
    label_row = 4
    ws.cell(header_row, col, "field_id").font = BOLD
    ws.cell(label_row, col, "科目").font = BOLD
    col += 1
    ticker_cols: Dict[str, Dict[str, int]] = {}
    for s in sorted(summaries, key=lambda x: (x["market"], x["id"])):
        key = f"{s['market']}:{s['id']}"
        ticker_cols[key] = {}
        for period in s.get("periods", []):
            ws.cell(header_row, col, f"{s['market']}-{s['id']}\n{period}").font = BOLD
            ws.cell(header_row, col).alignment = Alignment(wrap_text=True, horizontal="center")
            ticker_cols[key][period] = col
            col += 1

    row = 5
    for field_def in GLOBAL_FIELDS_V1:
        ws.cell(row, 1, field_def.field_id)
        ws.cell(row, 2, f"{field_def.label_zh} / {field_def.label_en}")
        for market, ticker, item in all_items:
            if item.field_id != field_def.field_id:
                continue
            key = f"{market}:{ticker}"
            c = ticker_cols.get(key, {}).get(item.period_end)
            if not c:
                continue
            if item.status == MatchStatus.SKIPPED:
                ws.cell(row, c, "—")
                ws.cell(row, c).fill = BLUE
            elif item.authoritative_value is not None:
                ws.cell(row, c, item.authoritative_value)
                ws.cell(row, c).fill = _status_fill(item.status)
        row += 1
    ws.freeze_panes = "C5"


def _write_items_sheet(
    ws,
    all_items: List[Tuple[str, str, ReconciliationItem]],
    *,
    with_delta: bool = False,
) -> None:
    registry = field_by_id()
    headers = [
        "Market", "Ticker", "Field", "科目", "Period", "推荐值", "XBRL", "PDF",
        "Status", "Trust", "Source",
    ]
    if with_delta:
        headers += ["Delta", "Delta%", "XBRL Tag", "PDF Label"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    for market, ticker, item in all_items:
        fd = registry.get(item.field_id)
        label = f"{fd.label_zh} / {fd.label_en}" if fd else item.field_id
        row_data = [
            market,
            ticker,
            item.field_id,
            label,
            item.period_end,
            item.authoritative_value,
            item.xbrl_value,
            item.pdf_value,
            item.status.value,
            item.trust_level.value,
            item.authoritative_source,
        ]
        if with_delta:
            row_data += [item.delta, item.delta_pct, item.xbrl_tag, item.pdf_label]
        ws.append(row_data)
        ws.cell(ws.max_row, 9).fill = _status_fill(item.status)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_warnings_sheet(ws, warnings: Dict[str, List[str]]) -> None:
    ws.append(["说明", "以下为 PDF 提取层诊断信息；若 XBRL校验=100% 则不影响已通过的对账结果"])
    ws.merge_cells("A1:D1")
    ws.append(["Market:Ticker", "警告", "说明"])
    _style_header(ws, 3, 3)
    for key, errs in sorted(warnings.items()):
        for err in errs:
            note = "已跳过对账" if any(
                k in err for k in ("缺少 XBRL 标签", "缺少报表页", "缺少报表网格", "PDF 未提取到")
            ) else ""
            ws.append([key, err, note])


def _build_workbook(
    summaries: List[dict],
    all_items: List[Tuple[str, str, ReconciliationItem]],
    warnings: Dict[str, List[str]],
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    ws_overview = wb.create_sheet("总览", 0)
    _write_overview_sheet(ws_overview, summaries)

    ws_sum = wb.create_sheet("校验汇总")
    _write_summary_sheet(ws_sum, summaries)

    if all_items:
        ws_rec = wb.create_sheet("推荐数据")
        _write_recommended_sheet(ws_rec, summaries, all_items)

        ws_data = wb.create_sheet("校验数据")
        _write_items_sheet(ws_data, all_items, with_delta=False)

        ws_recon = wb.create_sheet("对账明细")
        _write_items_sheet(ws_recon, all_items, with_delta=True)
    else:
        note = wb.create_sheet("说明")
        note.append(["快速模式：仅从 JSON 导出汇总，未包含财务数值。去掉 --fast 以跑完整管线。"])

    ws_warn = wb.create_sheet("提取警告")
    _write_warnings_sheet(ws_warn, warnings)

    for sheet in wb.worksheets:
        _auto_width(sheet)
    return wb


def export_verified_benchmark_excel(
    output_path: Path,
    *,
    markets: List[str],
    periods: int = 3,
    with_data: bool = True,
    us_report: Optional[Path] = US_REPORT,
    hk_report: Optional[Path] = HK_REPORT,
) -> str:
    markets = [m.upper() for m in markets]
    summaries: List[dict] = []
    all_items: List[Tuple[str, str, ReconciliationItem]] = []
    warnings: Dict[str, List[str]] = {}

    if with_data:
        if "US" in markets:
            print(f"US 10-K（{len(_us_templates())} 家）…")
            us_sum, us_items, us_warn = _collect_verified_rows(
                UsFinancialsService(), _us_templates(), "US", periods=periods
            )
            summaries.extend(us_sum)
            all_items.extend(us_items)
            warnings.update(us_warn)
        if "HK" in markets:
            print(f"港股（{len(_hk_templates())} 家）…")
            hk_sum, hk_items, hk_warn = _collect_verified_rows(
                HkFinancialsService(), _hk_templates(), "HK", periods=periods
            )
            summaries.extend(hk_sum)
            all_items.extend(hk_items)
            warnings.update(hk_warn)
    else:
        if "US" in markets:
            if not us_report or not us_report.exists():
                raise FileNotFoundError(f"缺少 US 报告: {us_report}")
            us_sum, us_warn = _summaries_from_report(us_report, "US")
            summaries.extend(us_sum)
            warnings.update(us_warn)
        if "HK" in markets:
            if not hk_report or not hk_report.exists():
                raise FileNotFoundError(f"缺少 HK 报告: {hk_report}")
            hk_sum, hk_warn = _summaries_from_report(hk_report, "HK")
            summaries.extend(hk_sum)
            warnings.update(hk_warn)

    wb = _build_workbook(summaries, all_items, warnings)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return str(output_path)


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="导出 verified benchmark Excel（US / HK / 合并）")
    parser.add_argument(
        "--market",
        choices=["us", "hk", "all"],
        default="all",
        help="导出市场范围（默认 all=40 家合并）",
    )
    parser.add_argument("--output", default=None, help="输出路径（默认按 market 选择）")
    parser.add_argument("--us-report", default=str(US_REPORT))
    parser.add_argument("--hk-report", default=str(HK_REPORT))
    parser.add_argument("--periods", type=int, default=3)
    parser.add_argument(
        "--fast",
        action="store_true",
        help="仅从 JSON 导出汇总/警告，不跑校验管线（无财务数值）",
    )
    args = parser.parse_args()

    if args.market == "us":
        markets = ["US"]
        default_out = US_OUTPUT
    elif args.market == "hk":
        markets = ["HK"]
        default_out = HK_OUTPUT
    else:
        markets = ["US", "HK"]
        default_out = MERGED_OUTPUT

    output = Path(args.output) if args.output else default_out
    path = export_verified_benchmark_excel(
        output,
        markets=markets,
        periods=args.periods,
        with_data=not args.fast,
        us_report=Path(args.us_report),
        hk_report=Path(args.hk_report),
    )
    print(f"Excel 已生成: {path}")
    print("Sheets: 总览 → 校验汇总 → 推荐数据 → 校验数据 → 对账明细 → 提取警告")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
