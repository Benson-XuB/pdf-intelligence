#!/usr/bin/env python3
"""把 10-K 基准结果导出为可读 Excel。"""

from __future__ import annotations

import json
import sys
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from backend.evaluation.financial_statement import (
    extract_ground_truth,
    extract_statement_pages_pdf,
    locate_statements,
    match_statement,
)
from backend.pipeline.orchestrator import PipelineOrchestrator
from scripts.run_financial_10k_benchmark import COMPANIES, REPORT_PATH

OUTPUT = ROOT / "data/outputs/financial_10k_benchmark.xlsx"
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BOLD = Font(bold=True)

STMT_ORDER = ("income", "balance", "cashflow")
STMT_CN = {"income": "利润表", "balance": "资产负债表", "cashflow": "现金流量表"}


def _safe_sheet_name(name: str) -> str:
    for ch in r"\/[]:*?":
        name = name.replace(ch, "_")
    return name[:31]


def export_excel() -> str:
    report = json.loads(REPORT_PATH.read_text()) if REPORT_PATH.exists() else {}
    orchestrator = PipelineOrchestrator()
    wb = Workbook()
    wb.remove(wb.active)

    # --- 汇总页 ---
    ws_sum = wb.create_sheet("汇总")
    ws_sum.append(["10-K 三大报表基准", "", ""])
    ws_sum.append(["整体准确率", f"{report.get('overall_accuracy', 0):.1%}", report.get("verdict", "")])
    ws_sum.append([])
    ws_sum.append(["公司", "平均准确率", "利润表", "资产负债表", "现金流量表"])
    for co in report.get("companies", []):
        stmt_acc = {s["statement"]: s["accuracy"] for s in co.get("statements", [])}
        ws_sum.append([
            f"{co['name']} ({co['id']})",
            f"{co['avg_accuracy']:.1%}",
            f"{stmt_acc.get('income', 0):.1%}" if "income" in stmt_acc else "—",
            f"{stmt_acc.get('balance', 0):.1%}" if "balance" in stmt_acc else "—",
            f"{stmt_acc.get('cashflow', 0):.1%}" if "cashflow" in stmt_acc else "—",
        ])
    for row in ws_sum.iter_rows(min_row=1, max_row=1):
        for c in row:
            c.font = BOLD

    benchmark_dir = ROOT / "tests/benchmark/financial_10k"
    benchmark_dir.mkdir(parents=True, exist_ok=True)

    for c in COMPANIES:
        pdf_path = ROOT / c["pdf"]
        if not pdf_path.exists():
            continue

        pages = locate_statements(str(pdf_path))
        if not pages:
            continue

        mini_path = benchmark_dir / f"{c['id']}_statements_only.pdf"
        extract_statement_pages_pdf(str(pdf_path), pages, str(mini_path))

        print(f"提取 {c['name']} ...")
        result = orchestrator.process(str(mini_path))
        ordered = [t for t in STMT_ORDER if t in pages]

        for mini_page, stype in enumerate(ordered):
            orig_page = pages[stype]
            truth = extract_ground_truth(str(pdf_path), stype, orig_page)
            page_tables = [t for t in result.tables if t.page_num == mini_page]
            df = page_tables[0].dataframe if page_tables else None
            source = page_tables[0].source if page_tables else "none"
            score = match_statement(truth, df if df is not None else __import__("pandas").DataFrame(), source=source)

            sheet = _safe_sheet_name(f"{c['id']}_{STMT_CN[stype]}")
            ws = wb.create_sheet(sheet)

            ws.cell(1, 1, f"{c['name']} — {STMT_CN[stype]} (p{orig_page + 1})").font = BOLD
            ws.cell(2, 1, f"准确率: {score.accuracy:.1%}")
            ws.cell(2, 2, f"来源: {source}")

            row = 4
            if df is not None and not df.empty:
                for c_idx, col in enumerate(df.columns, 1):
                    ws.cell(row, c_idx, str(col)).font = BOLD
                row += 1
                for _, data_row in df.iterrows():
                    for c_idx, val in enumerate(data_row, 1):
                        ws.cell(row, c_idx, str(val) if val is not None else "")
                    row += 1
            else:
                ws.cell(row, 1, "（未提取到表格）")
                row += 2

            row += 1
            ws.cell(row, 1, "逐项对比（官方 vs 提取）").font = BOLD
            row += 1
            ws.cell(row, 1, "行项").font = BOLD
            ws.cell(row, 2, "官方").font = BOLD
            ws.cell(row, 3, "提取").font = BOLD
            ws.cell(row, 4, "结果").font = BOLD
            row += 1
            for item in score.items:
                ws.cell(row, 1, item.key)
                ws.cell(row, 2, str(item.expected))
                ws.cell(row, 3, str(item.actual))
                status_cell = ws.cell(row, 4, "OK" if item.matched else "MISS")
                status_cell.fill = GREEN if item.matched else RED
                row += 1

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return str(OUTPUT)


if __name__ == "__main__":
    t0 = time.time()
    path = export_excel()
    print(f"\nExcel 已生成: {path}")
    print(f"耗时: {time.time() - t0:.0f}s")
