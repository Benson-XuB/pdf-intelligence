from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.export.formula_excel import FormulaExcelExporter
from backend.global_schema.models import CompanyFinancials, FieldValue, ValueScale
from backend.markets.us.statement_grid_extractor import StatementGrid
from backend.validation.identity_models import IdentityCheckItem, IdentityReport
from backend.validation.reconciliation import (
    FinancialReconciler,
    MatchStatus,
    ReconciliationReport,
)


def _field(field_id: str, period: str, value: float, source: str = "xbrl") -> FieldValue:
    return FieldValue(
        field_id=field_id,
        period_end=period,
        fiscal_year=int(period[:4]),
        value=value,
        scale=ValueScale.MILLIONS,
        standard="US-GAAP",
        source=source,
        source_tag=source,
    )


def test_formula_excel_writes_sum_and_validation(tmp_path):
    periods = ["2024-12-31", "2023-12-31"]
    xbrl = CompanyFinancials(
        ticker="DEMO",
        company_name="Demo Inc.",
        market="US",
        cik="0000000000",
        standard="US-GAAP",
        periods=periods,
        values=[
            _field("total_assets", periods[0], 1000.0),
            _field("total_liabilities", periods[0], 600.0),
            _field("total_equity", periods[0], 400.0),
            _field("revenue", periods[0], 500.0),
        ],
    )
    pdf = CompanyFinancials(
        ticker="DEMO",
        company_name="Demo Inc.",
        market="US",
        cik="0000000000",
        standard="US-GAAP",
        periods=periods,
        values=[
            _field("total_assets", periods[0], 1000.0, source="html_grid"),
            _field("total_liabilities", periods[0], 600.0, source="html_grid"),
            _field("total_equity", periods[0], 400.0, source="html_grid"),
            _field("revenue", periods[0], 500.0, source="html_grid"),
        ],
    )
    reconciliation = FinancialReconciler().reconcile(xbrl, pdf)
    identity = IdentityReport(
        standard="US-GAAP",
        items=[
            IdentityCheckItem(
                rule_id="balance_equation",
                label="Total Assets = Total Liabilities + Total Equity",
                period_end=periods[0],
                passed=True,
                lhs_value=1000.0,
                rhs_value=1000.0,
                delta=0.0,
            )
        ],
    )
    grids = {
        "balance": StatementGrid(
            statement_type="balance",
            period_ends=[periods[0]],
                rows=[
                    ("Cash", ["1000"]),
                    ("Property", ["3000"]),
                    ("Total assets", ["4000"]),
                ],
        )
    }

    out = tmp_path / "demo_formula_model.xlsx"
    FormulaExcelExporter().export(
        ticker="DEMO",
        company_name="Demo Inc.",
        standard="US-GAAP",
        reconciliation=reconciliation,
        identity_report=identity,
        statement_grids=grids,
        authoritative=xbrl,
        output_path=str(out),
    )

    wb = load_workbook(out, data_only=False)
    assert "Model Summary" in wb.sheetnames
    assert "Validation" in wb.sheetnames
    assert "Model - Balance" in wb.sheetnames

    balance = wb["Model - Balance"]
    total_cell = None
    for row in range(1, balance.max_row + 1):
        if balance.cell(row=row, column=1).value == "Total assets":
            total_cell = balance.cell(row=row, column=2)
            break
    assert total_cell is not None
    assert str(total_cell.value).startswith("=SUM(")

    validation = wb["Validation"]
    status_formula = validation.cell(row=4, column=7).value
    assert isinstance(status_formula, str)
    assert status_formula.startswith("=IF(")
    assert "PASS" in status_formula
